import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
import logging

from typing import Dict, Any, Optional

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 确保data目录存在
data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
if not os.path.exists(data_dir):
    os.makedirs(data_dir)
    logger.info(f"创建数据目录: {data_dir}")

# 消息存储Excel文件路径 - 使用绝对路径
EXCEL_FILE_PATH = os.path.join(data_dir, "messages.xlsx")
logger.info(f"Excel文件路径: {EXCEL_FILE_PATH}")

class MessageService:
    
    @staticmethod
    def parse_xml_message(xml_data: str) -> Optional[Dict[str, Any]]:
        """
        解析微信公众号推送的XML消息
    
        参数:
         xml_data: XML格式的字符串数据
        
        返回:
         Dict: 解析后的消息字典，包含关键字段
        """
        try:
            # 解析XML数据
            root = ET.fromstring(xml_data)
            
            # 提取基础消息字段
            message_data = {
                'to_user_name': root.find('ToUserName').text if root.find('ToUserName') is not None else '',
                'from_user_name': root.find('FromUserName').text if root.find('FromUserName') is not None else '',
                'create_time': root.find('CreateTime').text if root.find('CreateTime') is not None else '',
                'msg_type': root.find('MsgType').text if root.find('MsgType') is not None else '',
                'msg_id': root.find('MsgId').text if root.find('MsgId') is not None else '',
            }
            
            # 添加格式化时间字段
            if message_data['create_time']:
                try:
                    timestamp = int(message_data['create_time'])
                    message_data['create_time_formatted'] = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    message_data['create_time_formatted'] = "Invalid time"
            else:
                message_data['create_time_formatted'] = ""
            
            # 根据消息类型提取特定内容
            msg_type = message_data['msg_type']
            
            if msg_type == 'text':
                # 文本消息
                message_data['content'] = root.find('Content').text if root.find('Content') is not None else ''
            elif msg_type == 'image':
                # 图片消息
                message_data['pic_url'] = root.find('PicUrl').text if root.find('PicUrl') is not None else ''
                message_data['media_id'] = root.find('MediaId').text if root.find('MediaId') is not None else ''
            elif msg_type == 'voice':
                # 语音消息
                message_data['media_id'] = root.find('MediaId').text if root.find('MediaId') is not None else ''
                message_data['format'] = root.find('Format').text if root.find('Format') is not None else ''
            elif msg_type == 'video' or msg_type == 'shortvideo':
                # 视频消息
                message_data['media_id'] = root.find('MediaId').text if root.find('MediaId') is not None else ''
                message_data['thumb_media_id'] = root.find('ThumbMediaId').text if root.find('ThumbMediaId') is not None else ''
            elif msg_type == 'location':
                # 位置消息
                message_data['location_x'] = root.find('Location_X').text if root.find('Location_X') is not None else ''
                message_data['location_y'] = root.find('Location_Y').text if root.find('Location_Y') is not None else ''
                message_data['scale'] = root.find('Scale').text if root.find('Scale') is not None else ''
                message_data['label'] = root.find('Label').text if root.find('Label') is not None else ''
            elif msg_type == 'link':
                # 链接消息
                message_data['title'] = root.find('Title').text if root.find('Title') is not None else ''
                message_data['description'] = root.find('Description').text if root.find('Description') is not None else ''
                message_data['url'] = root.find('Url').text if root.find('Url') is not None else ''
            elif msg_type == 'event':
                # 事件消息
                message_data['event'] = root.find('Event').text if root.find('Event') is not None else ''
                message_data['event_key'] = root.find('EventKey').text if root.find('EventKey') is not None else ''
            
            return message_data
            
        except Exception as e:
            print(f"解析XML消息时出错: {e}")
            return None

    @staticmethod
    def save_message_to_excel(message):
        """
        将解析后的消息数据追加保存到Excel文件
        :param message: 解析后的消息字典
        :return: 保存结果 (True/False)
        """
        try:
            logger.info(f"开始保存消息到Excel，消息ID: {message.get('msg_id', '未知')}")
            
            # 检查文件是否存在，如果不存在则创建并添加表头
            if not os.path.exists(EXCEL_FILE_PATH):
                logger.info(f"Excel文件不存在，创建新文件: {EXCEL_FILE_PATH}")
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "消息记录"
                # 添加表头
                headers = [
                    "消息ID", "接收时间", "格式化时间", "消息类型", 
                    "发送者OpenID", "接收者ID", "内容", "图片URL", 
                    "媒体ID", "格式", "缩略图媒体ID"
                ]
                sheet.append(headers)
                # 设置表头居中对齐
                for col in sheet.columns:
                    for cell in col:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                workbook.save(EXCEL_FILE_PATH)
                logger.info("Excel文件创建成功")

            # 打开现有文件并追加数据
            logger.info("打开Excel文件追加数据")
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active

            # 准备要写入的数据行 - 使用与parse_xml_message一致的字段名
            row_data = [
                message.get("msg_id", ""),
                message.get("create_time", ""),
                message.get("create_time_formatted", ""),
                message.get("msg_type", ""),
                message.get("from_user_name", ""),
                message.get("to_user_name", ""),
                message.get("content", ""),
                message.get("pic_url", ""),
                message.get("media_id", ""),
                message.get("format", ""),
                message.get("thumb_media_id", "")
            ]

            sheet.append(row_data)
            workbook.save(EXCEL_FILE_PATH)
            logger.info(f"消息保存成功，文件路径: {EXCEL_FILE_PATH}")
            return True
        except Exception as e:
            logger.error(f"保存消息到Excel失败: {str(e)}", exc_info=True)
            return False

    @staticmethod
    def get_all_messages():
        """
        从Excel文件中获取所有消息记录
        :return: 消息记录列表
        """
        try:
            logger.info("开始获取所有消息记录")
            
            if not os.path.exists(EXCEL_FILE_PATH):
                logger.info(f"Excel文件不存在: {EXCEL_FILE_PATH}")
                return []

            logger.info(f"打开Excel文件读取数据: {EXCEL_FILE_PATH}")
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active
            messages = []

            # 获取表头
            headers = [cell.value for cell in sheet[1]]
            logger.info(f"获取表头成功: {headers}")

            # 遍历数据行
            for row in sheet.iter_rows(min_row=2, values_only=True):
                message_dict = {}
                for header, value in zip(headers, row):
                    message_dict[header] = value
                messages.append(message_dict)

            logger.info(f"读取消息记录成功，共 {len(messages)} 条")
            return messages
        except Exception as e:
            logger.error(f"读取消息记录失败: {str(e)}", exc_info=True)
            return []